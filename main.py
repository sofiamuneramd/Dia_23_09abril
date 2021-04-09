# Importamos una libreria para tratar archivos .xlsx y c

from openpyxl import load_workbook
import xlsxwriter
import numpy as np

# Creamos unas variables que contienen el libro y la hoja que deseamos leer

file_path='Libro1.xlsx'
SHEET='Hoja1'

# Con el load_workbook abrimos el file_path solo como lectura

workbook=load_workbook(file_path,read_only=True)

# Luego abrimos la hoja que vamos a leer

sheet=workbook[SHEET]

# Quiero ver el contenido de la hoja para saber que valores estaré usando 

for row in sheet.iter_rows():
  print(row[0].value)
  print(row[1].value)
  print(row[2].value)

# Le asigno a cada celda una variable para poder usar su valor luego

a1 = sheet['A1'].value
a2 = sheet['A2'].value
a3 = sheet['A3'].value
b1 = sheet['B1'].value
b2 = sheet['B2'].value
b3 = sheet['B3'].value
c1 = sheet['C1'].value
c2 = sheet['C2'].value
c3 = sheet['C3'].value

# Con estas variables vamos a crear una matriz con numpy

m1 = np.array([[a1,a2,a3]])
m2 = np.array([[b1,b2,b3]])
m4 = np.ones((1,3))
print(m4)
m5 = (np.array([[c1,c2,c3]]))**2
print(m5)

# Realizamos operaciones entre matrices

m3=m1*m2
print(m3)
m6=m5+m4


# Leemos los valores de las entradas de la matriz

m1a=m1[0,0]
m1b=m1[0,1]
m1c=m1[0,2]
m3a=m3[0,0]
m3b=m3[0,1]
m3c=m3[0,2]
m6a=m6[0,0]
m6b=m6[0,1]
m6c=m6[0,2]


# Creamos un nuevo libro y hoja

workbook = xlsxwriter.Workbook("hoja1.xlsx")
worksheet = workbook.add_worksheet()

# Creamos una matriz que será introducida a la hoja. No se hace con numpy ya que estas matrices no son iterables

m1=[[m1a,m1b,m1c]]

row = 0

# Con un ciclo for escribimos en la hoja

for col, data in enumerate(m1):
  worksheet.write_column(row, col, data)

m3=[[m3a,m3b,m3c]]

col=2

for row, data in enumerate(m3):
  worksheet.write_row(col,row,data)

m6=[[m6a,m6b,m6c]]

row = 1

# Con un ciclo for escribimos en la hoja

for col, data in enumerate(m6):
  worksheet.write_column(row, col, data)


# Cerramos el archivo para que se pueda guardar y crear

workbook.close()

'''


workbook=load_workbook('Libro2.xlsx',read_only=True)

# Luego abrimos la hoja que vamos a leer

sheet=workbook['Hoja2']
'''






